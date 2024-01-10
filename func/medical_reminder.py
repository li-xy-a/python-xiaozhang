# coding=utf-8
import openpyxl
from datetime import datetime
import config.global_params as gp

"""
    医疗期提醒 函数
"""


def get_expire():
    file_path = gp.sick_leave_execl_path
    now = datetime.now()
    # 返回值为二维数组  里边每条的数据为 ：序号， 姓名，医疗期
    # 读取execl
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.worksheets[0]
    rows = list(sheet.rows)
    # 医疗期提醒sheet
    inform_sheet = wb[gp.sick_leave_execl_sheet_name]
    inform_values = inform_sheet.iter_cols(min_col=1, max_col=1, values_only=True)
    inform_list = tuple(inform_values)[0]
    res = []
    # 读取数据
    for index in range(2, len(rows)):
        row = rows[index]
        # 如果已经在医疗提醒sheet里不处理
        if row[0].value in inform_list:
            continue
        # 获取当前行的结束时间
        end_date = row[13].value
        days_difference = (end_date - now).days
        if 30 >= days_difference >= 1:
            res.append([row[0].value, row[1].value, row[2].value, end_date, '未提醒'])
    input_execl_date(file_path, res)
    wb.close()


"""
    将需要提醒的人输出到execl中的指定sheet中去
"""


def input_execl_date(file_path, data):
    if data is not None and len(data) > 0:
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb[gp.sick_leave_execl_sheet_name]
            for row in data:
                sheet.append(row)
            wb.save(file_path)
        except Exception as e:
            print(f"发生错误：{e}")


"""
    获取医疗通知列表的列表
"""


def get_list():
    file_path = gp.sick_leave_execl_path
    list_res = []
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[gp.sick_leave_execl_sheet_name]
        for row in sheet.iter_rows(values_only=True):
            list_res.append(row)
        # 保留第一条数据
        first = list_res[0]
        list_res.remove(first)
        list_res = sorted(list_res, key=lambda x: (x[4], x[3],), reverse=True)
        list_res.insert(0, first)
    except Exception as e:
        print(f"Error: {e}")
    return list_res


"""
    更改状态为已提醒
"""


def remind(id):
    file_path = gp.sick_leave_execl_path
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[gp.sick_leave_execl_sheet_name]

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == id:
                sheet.cell(row=row_index, column=5, value='已提醒')

        wb.save(file_path)
    except Exception as e:
        print(f"Error: {e}")


if __name__ == '__main__':
    print(get_list())
